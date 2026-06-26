"""
Kingdom Legacy — Backend de Procesamiento Financiero
=====================================================
API REST para parsear documentos SII y Excel de ventas,
calcular indicadores de salud financiera y proyecciones.

Autor: Kingdom Legacy
"""

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import tempfile, os, re, shutil
from typing import Optional
import subprocess
from PIL import Image
import openpyxl
import pdfplumber

app = FastAPI(
    title="Kingdom Legacy — API Diagnóstico Financiero",
    description="Procesa documentos SII y Excel para generar diagnósticos financieros",
    version="1.0.0"
)

# CORS: permite que el dashboard HTML llame a esta API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── UTILIDADES ────────────────────────────────────────────────────────────────

def clean_num(s: str) -> int:
    """Limpia string numérico chileno (1.234.567) a int."""
    try:
        return int(re.sub(r'[.,\s]', '', str(s)))
    except:
        return 0

MONTH_MAP = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
    'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
    'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
}

def detect_period(text: str) -> Optional[str]:
    """Extrae el período AAAA-MM del texto OCR de un F29."""
    # Campo PERIODO AAAAMM
    m = re.search(r'PERIODO[^\d]*(\d{6})', text, re.IGNORECASE)
    if m:
        raw = m.group(1)
        return f"{raw[:4]}-{raw[4:]}"
    # Nombre de mes + año
    for name, num in MONTH_MAP.items():
        m2 = re.search(rf'\b{name}\b[^\d]*(\d{{4}})', text, re.IGNORECASE)
        if m2:
            return f"{m2.group(1)}-{str(num).zfill(2)}"
    return None

def parse_f29_fields(crop_text: str) -> dict:
    """
    Extrae campos clave del F29 desde texto OCR de la zona de tabla.
    Campos: base imponible, débitos IVA, créditos IVA, IVA a pagar, PPM.
    """
    def find(patterns):
        for p in patterns:
            m = re.search(p, crop_text, re.IGNORECASE)
            if m:
                return clean_num(m.group(1))
        return 0

    return {
        "base_imponible": find([
            r'BASE\s+IMPONIBLE\s+([\d.,]+)',
            r'\b563\b[^\d]+([\d.,]+)'
        ]),
        "debitos_iva": find([
            r'DEBITOS\s+FACTURAS\s+EMITIDAS\s+([\d.,]+)',
            r'TOTAL\s+D[EÉ]BITOS?\s+([\d.,]+)',
            r'\b336\b[^\d]+([\d.,]+)'
        ]),
        "creditos_iva": find([
            r'TOTAL\s+CR[EÉ]DITOS?\s+([\d.,]+)',
            r'\b337\b[^\d]+([\d.,]+)'
        ]),
        "iva_pagar": find([
            r'TOTAL\s+A\s+PAGAR\s+DENTRO\s+DEL\s+PLAZO\s+LEGAL\s+([\d.,]+)',
            r'TOTAL\s+DETERMINADO\s+([\d.,]+)',
            r'\b547\b[^\d]+([\d.,]+)'
        ]),
        "ppm": find([
            r'PPM\s+NETO\s+DETERMINADO\s+([\d.,]+)',
            r'\b052\b[^\d]+([\d.,]+)'
        ]),
    }

def ocr_pdf_page(pdf_path: str, page_num: int) -> tuple[str, str]:
    """
    OCR de una página del PDF.
    Retorna (texto_completo, texto_tabla_recortada).
    """
    out_prefix = f'/tmp/kl_ocr_{page_num}'
    subprocess.run(
        ['pdftoppm', '-r', '300', '-f', str(page_num), '-l', str(page_num),
         '-png', pdf_path, out_prefix],
        check=True, capture_output=True
    )
    files = sorted([f for f in os.listdir('/tmp') if f.startswith(f'kl_ocr_{page_num}')])
    if not files:
        return "", ""

    img_path = f'/tmp/{files[-1]}'
    # OCR página completa
    r_full = subprocess.run(
        ['tesseract', img_path, 'stdout', '--psm', '6'],
        capture_output=True, text=True
    ).stdout

    # OCR zona de tabla (recorte vertical 28%–84%)
    img = Image.open(img_path)
    w, h = img.size
    cropped = img.crop((0, int(h * 0.26), w, int(h * 0.84)))
    crop_path = f'/tmp/kl_crop_{page_num}.png'
    cropped.save(crop_path)
    r_crop = subprocess.run(
        ['tesseract', crop_path, 'stdout', '--psm', '6'],
        capture_output=True, text=True
    ).stdout

    # Limpieza
    for f in files:
        try: os.remove(f'/tmp/{f}')
        except: pass
    try: os.remove(crop_path)
    except: pass

    return r_full, r_crop


# ── ENDPOINTS ─────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {
        "sistema": "Kingdom Legacy — API Diagnóstico Financiero",
        "version": "1.0.0",
        "endpoints": {
            "POST /upload/f29": "Sube PDF de Carpeta Tributaria SII",
            "POST /upload/ventas": "Sube Excel de libro de ventas",
            "GET /diagnostico/{empresa_id}": "Obtiene diagnóstico completo",
        }
    }


@app.post("/upload/f29")
async def upload_f29(file: UploadFile = File(...)):
    """
    Recibe el PDF de la Carpeta Tributaria SII.
    Extrae todos los F29 disponibles y retorna los datos por período.
    """
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(400, "Solo se aceptan archivos PDF")

    # Guardar temporalmente
    tmp = tempfile.mktemp(suffix='.pdf')
    try:
        with open(tmp, 'wb') as f:
            content = await file.read()
            f.write(content)

        # Contar páginas
        with pdfplumber.open(tmp) as pdf:
            total_pages = len(pdf.pages)

        resultados = []
        errores = []

        for pg in range(1, total_pages + 1):
            try:
                full_text, crop_text = ocr_pdf_page(tmp, pg)

                # Solo procesar páginas que sean F29
                if 'FORMULARIO 29' not in full_text.upper():
                    continue

                period = detect_period(full_text)
                if not period:
                    continue

                fields = parse_f29_fields(crop_text)
                resultados.append({
                    "period": period,
                    **fields,
                    "pagina_pdf": pg
                })

            except Exception as e:
                errores.append({"pagina": pg, "error": str(e)})

        # Ordenar cronológicamente
        resultados.sort(key=lambda x: x['period'])

        return {
            "ok": True,
            "archivo": file.filename,
            "total_paginas": total_pages,
            "periodos_extraidos": len(resultados),
            "errores": len(errores),
            "datos": resultados
        }

    finally:
        try: os.remove(tmp)
        except: pass


@app.post("/upload/ventas")
async def upload_ventas(file: UploadFile = File(...)):
    """
    Recibe Excel de libro de ventas (formato REDITAL o similar).
    Extrae ventas totales, netas, costo, rentabilidad por período.
    """
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Solo se aceptan archivos Excel (.xlsx, .xls)")

    tmp = tempfile.mktemp(suffix='.xlsx')
    try:
        with open(tmp, 'wb') as f:
            content = await file.read()
            f.write(content)

        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
        resultado = {
            "archivo": file.filename,
            "hojas": wb.sheetnames,
            "resumen": {},
            "ventas_detalle": [],
            "por_forma_pago": {}
        }

        # Hoja de ventas
        hoja_ventas = None
        for name in wb.sheetnames:
            if 'VENTA' in name.upper():
                hoja_ventas = wb[name]
                break

        if hoja_ventas:
            rows = list(hoja_ventas.iter_rows(max_row=8, values_only=True))

            # Extraer resumen (filas 1-6)
            for row in rows:
                if not row: continue
                for i, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        val = row[i+1] if i+1 < len(row) else None
                        if 'VENTAS TOTALES' in str(cell).upper() and val:
                            resultado['resumen']['ventas_totales'] = val
                        elif 'VENTAS NETAS' in str(cell).upper() and val:
                            resultado['resumen']['ventas_netas'] = val
                        elif 'COSTO DE VENTA' in str(cell).upper() and val:
                            resultado['resumen']['costo_venta'] = val
                        elif 'RENTABILIDAD NETA' in str(cell).upper() and val and '% ' not in str(cell):
                            resultado['resumen']['utilidad_neta'] = val
                        elif '% RENTABILIDAD' in str(cell).upper() and val:
                            resultado['resumen']['rentabilidad_pct'] = round(float(val) * 100, 2) if val else 0

                    # Forma de pago
                    if cell in ['TRANSFERENCIA ', 'TRANSFERENCIA', 'TRANSBANK', 'FLOW', 'EFECTIVO ', 'EFECTIVO']:
                        val = row[i+1] if i+1 < len(row) else None
                        if val and isinstance(val, (int, float)):
                            resultado['por_forma_pago'][str(cell).strip()] = val

            # Detectar período del archivo por fechas en columna B
            fechas = []
            for row in hoja_ventas.iter_rows(min_row=9, max_row=100, values_only=True):
                if row and row[1] and hasattr(row[1], 'year'):
                    fechas.append(row[1])
            if fechas:
                from datetime import datetime
                año = fechas[0].year
                mes = fechas[0].month
                resultado['resumen']['period'] = f"{año}-{str(mes).zfill(2)}"

        return {"ok": True, **resultado}

    finally:
        try: os.remove(tmp)
        except: pass


@app.get("/health")
def health():
    """Verificar que la API está corriendo."""
    return {"status": "ok", "sistema": "Kingdom Legacy API"}
